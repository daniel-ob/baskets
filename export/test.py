from io import BytesIO

from django.contrib.auth import get_user_model
from django.urls import reverse, reverse_lazy
import openpyxl

from baskets.models import Producer
from baskets.tests.common import BasketsTestCase

WORKSHEET_NAME_MAX_LENGTH = 31


class TestDeliveryExport(BasketsTestCase):
    def test_success(self):
        self.u2.username = (
            "username_with_more_than_31_chars_that_must_be_cut_on_sheet_name"
        )
        self.u2.save()

        d = self.d2
        # staff member required for export
        self.u1.is_staff = True
        self.u1.save()
        self.client.force_login(self.u1)
        response = self.client.get(reverse("delivery_export", args=[d.id]))
        self.assertEqual(response.status_code, 200)

        self.assertIn(
            f"filename={str(d.date)}", response.headers["Content-Disposition"]
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertEqual(len(wb.worksheets), d.orders.count())  # one sheet per d.order
        sheets = {sheet.title: sheet for sheet in wb.worksheets}
        for order in d.orders.all():
            cut_username = order.user.username[:WORKSHEET_NAME_MAX_LENGTH]
            self.assertIn(cut_username, sheets.keys())
            sheet = sheets[cut_username]
            rows = list(sheet.iter_rows(values_only=True))
            order_amount = rows[-1][-1]
            self.assertEqual(f"{order_amount:.2f}", f"{order.amount:.2f}")

    def test_not_staff(self):
        url = reverse("delivery_export", args=[self.d2.id])
        response = self.client.get(url)
        self.assertRedirects(response, f"{reverse('admin:login')}?next={url}")


class TestOrderExport(BasketsTestCase):
    url = reverse_lazy("order_export")

    def test_success(self):
        # staff member required for export
        self.u1.is_staff = True
        self.u1.save()
        self.client.force_login(self.u1)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        self.assertIn(
            "filename=order_export.xlsx", response.headers["Content-Disposition"]
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertEqual(len(wb.worksheets), 1)
        rows = list(wb.worksheets[0].iter_rows(values_only=True))
        self.assertEqual(
            len(rows), get_user_model().objects.count() + 1
        )  # one row per user +1 for header

    def test_not_staff(self):
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{reverse('admin:login')}?next={self.url}")


class TestProducerExport(BasketsTestCase):
    url = reverse_lazy("producer_export")

    def test_success(self):
        self.producer1.name = (
            "A long name with more than 31 chars"  # name must be cut on sheet name
        )
        self.producer1.save()

        # staff member required for export
        self.u1.is_staff = True
        self.u1.save()
        self.client.force_login(self.u1)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        self.assertIn(
            "filename=producer_export.xlsx", response.headers["Content-Disposition"]
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertEqual(
            len(wb.worksheets), Producer.objects.count()
        )  # one sheet per Producer, inactive ones are also exported
        sheets = {sheet.title: sheet for sheet in wb.worksheets}
        for producer, product_count in [
            (self.producer1, 2),
            (self.producer4, 1),  # inactive producer with 1 inactive product
        ]:
            sheet = sheets[producer.name[:WORKSHEET_NAME_MAX_LENGTH]]
            rows = list(sheet.iter_rows(values_only=True))
            self.assertEqual(
                len(rows), product_count + 1
            )  # one row per product +1 for header

    def test_not_staff(self):
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{reverse('admin:login')}?next={self.url}")
