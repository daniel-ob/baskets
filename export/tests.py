from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse, reverse_lazy
from openpyxl import load_workbook

from baskets.tests.common import (
    create_user,
    create_closed_delivery,
    create_order_item,
    create_producer,
    create_product,
)

WORKSHEET_NAME_MAX_LENGTH = 31


class TestDeliveryExport(TestCase):
    def test_success(self):
        user = create_user()
        user.username = (
            "username_with_more_than_31_chars_that_must_be_cut_on_sheet_name"
        )
        user.save()
        d = create_closed_delivery()
        create_order_item(delivery=d, user=user)
        # order from another user
        create_order_item(delivery=d)

        self.client.force_login(create_user(is_staff=True))
        response = self.client.get(reverse("delivery_export", args=[d.id]))
        self.assertEqual(response.status_code, 200)

        self.assertIn(
            f"filename={str(d.date)}", response.headers["Content-Disposition"]
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(len(wb.worksheets), d.orders.count())  # one sheet per d.order
        sheets = {sheet.title: sheet for sheet in wb.worksheets}
        for order in d.orders.all():
            cut_username = order.user.username[:WORKSHEET_NAME_MAX_LENGTH]
            self.assertIn(cut_username, sheets.keys())
            sheet = sheets[cut_username]
            rows = list(sheet.iter_rows(values_only=True))
            order_amount = rows[-1][-1]
            self.assertEqual(f"{order_amount:.2f}", f"{order.amount:.2f}")

    def test_not_staff(self):
        url = reverse("delivery_export", args=[create_closed_delivery().id])
        response = self.client.get(url)
        self.assertRedirects(response, f"{reverse('admin:login')}?next={url}")


class TestOrderExport(TestCase):
    url = reverse_lazy("order_export")

    def test_success(self):
        users = [create_user() for _ in range(3)]
        # create deliveries for 2 different months
        deliveries = [create_closed_delivery() for _ in range(2)]
        deliveries[1].date += timedelta(days=31)
        deliveries[1].save()

        # staff member required for export
        self.client.force_login(create_user(is_staff=True))
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        self.assertIn(
            "filename=order_export.xlsx", response.headers["Content-Disposition"]
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(len(wb.worksheets), 1)
        rows = list(wb.worksheets[0].iter_rows())
        self.assertEqual(
            len(rows), get_user_model().objects.count() + 1
        )  # one row per user (staff included) +1 for header
        cols = list(wb.worksheets[0].iter_cols())
        self.assertEqual(
            len(cols), len(deliveries) + 1
        )  # one row per month +1 for usernames

    def test_not_staff(self):
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{reverse('admin:login')}?next={self.url}")


class TestProducerExport(TestCase):
    url = reverse_lazy("producer_export")

    def test_success(self):
        producers = [create_producer() for _ in range(3)]
        [create_product(producer=producer) for _ in range(4) for producer in producers]
        producers[
            0
        ].name = "Long name with more than 31 chars must be cut on sheet name"
        producers[0].save()
        # not active producer and products
        producers[1].is_active = False
        producers[1].save()
        assert not producers[1].products.first().is_active

        # staff member required for export
        self.client.force_login(create_user(is_staff=True))
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        self.assertIn(
            "filename=producer_export.xlsx", response.headers["Content-Disposition"]
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(
            len(wb.worksheets), len(producers)
        )  # one sheet per Producer, inactive ones are also exported
        sheets = {sheet.title: sheet for sheet in wb.worksheets}
        for producer in producers:
            sheet = sheets[producer.name[:WORKSHEET_NAME_MAX_LENGTH]]
            rows = list(sheet.iter_rows(values_only=True))
            self.assertEqual(
                len(rows), producer.products.count() + 1
            )  # one row per product (inactive ones included) +1 for header

    def test_not_staff(self):
        response = self.client.get(self.url)
        self.assertRedirects(response, f"{reverse('admin:login')}?next={self.url}")
